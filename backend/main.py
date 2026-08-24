from __future__ import annotations

import asyncio
import contextlib
import json
import math
import os
import uuid
from collections import OrderedDict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from threading import Lock
from time import monotonic
from typing import Any
from urllib.parse import quote_plus
from urllib.request import Request as UrlRequest, urlopen

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from pydantic import BaseModel, Field
from sqlalchemy import JSON, BigInteger, DateTime, LargeBinary, String, create_engine, func, select, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker

from processor import generate_mock_dashboard, process_excel_file


APP_NAME = os.getenv("APP_NAME", "CJ WMS Dashboard API")
USE_MOCK = os.getenv("USE_MOCK", "true").lower() == "true"
UPLOAD_ROOT = Path(os.getenv("UPLOAD_ROOT", "../uploads")).resolve()
SYSTEM_ID = str(uuid.uuid4())
workspace_event_subscribers: set[asyncio.Queue[tuple[str, str]]] = set()
EXCEL_CACHE_TTL_SECONDS = int(os.getenv("EXCEL_CACHE_TTL_SECONDS", "600"))
EXCEL_CACHE_MAX_SHEETS = int(os.getenv("EXCEL_CACHE_MAX_SHEETS", "12"))
FILE_SERVER_SYNC_ENABLED = os.getenv("FILE_SERVER_SYNC_ENABLED", "false").lower() == "true"
FILE_SERVER_PATH = os.getenv("FILE_SERVER_PATH", r"\\10.84.194.51\CJWMSDashboard")
FILE_SERVER_SYNC_SECONDS = max(60, int(os.getenv("FILE_SERVER_SYNC_SECONDS", "1800")))
FILE_SERVER_MAX_FILE_BYTES = int(os.getenv("FILE_SERVER_MAX_FILE_MB", "25")) * 1024 * 1024
FILE_SERVER_MIRROR_URL = os.getenv("FILE_SERVER_MIRROR_URL", "").strip()
FILE_SERVER_MIRROR_TOKEN = os.getenv("FILE_SERVER_MIRROR_TOKEN", "")
excel_sheet_cache: OrderedDict[tuple[str, str], tuple[float, str, list[list[Any]], list[list[str]]]] = OrderedDict()
excel_cache_lock = Lock()
excel_sheet_locks: dict[tuple[str, str], Lock] = {}
file_server_sync_lock = Lock()
file_server_sync_task: asyncio.Task[None] | None = None
file_server_sync_status: dict[str, Any] = {
    "enabled": FILE_SERVER_SYNC_ENABLED,
    "path": FILE_SERVER_PATH,
    "interval_seconds": FILE_SERVER_SYNC_SECONDS,
    "state": "disabled" if not FILE_SERVER_SYNC_ENABLED else "waiting",
    "last_checked_at": None,
    "last_synced_at": None,
    "latest_filename": None,
    "upload_id": None,
    "message": "Automatic sync is disabled" if not FILE_SERVER_SYNC_ENABLED else "Waiting for first scan",
}


def broadcast_workspace_event(event_name: str, data: str) -> None:
    for event_queue in tuple(workspace_event_subscribers):
        try:
            event_queue.put_nowait((event_name, data))
        except asyncio.QueueFull:
            pass
DATABASE_URL = os.getenv("DATABASE_URL", "")
MYSQL_HOST = os.getenv("MYSQL_HOST", "localhost")
MYSQL_PORT = os.getenv("MYSQL_PORT", "3306")
MYSQL_DATABASE = os.getenv("MYSQL_DATABASE", "cj_wms_dashboard")
MYSQL_USER = os.getenv("MYSQL_USER", "cj_wms_user")
MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD", "change_me")

app = FastAPI(title=APP_NAME)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class LoginRequest(BaseModel):
    username: str
    password: str


class DashboardSettings(BaseModel):
    refresh_seconds: int = 60
    theme: str = "light"
    show_inbound: bool = True
    show_pick: bool = True
    show_outbound: bool = True


class DashboardWidgetItem(BaseModel):
    id: str
    type: str
    slot: int
    label: str


class DashboardWidget(BaseModel):
    id: str
    title: str
    size: str
    items: list[DashboardWidgetItem] = Field(default_factory=list)


class DashboardWidgetLayout(BaseModel):
    widgets: list[DashboardWidget]


class BarDataItem(BaseModel):
    id: str
    label: str
    cell: str
    markerCell: str | None = None


class StackBarCategory(BaseModel):
    id: str
    label: str


class StackBarSeries(BaseModel):
    id: str
    label: str
    cells: list[str] = Field(default_factory=list)


class LineAnnotation(BaseModel):
    id: str
    label: str
    axis: str = "y"
    value: str
    color: str = "#e42f44"


class GradientStop(BaseModel):
    id: str
    color: str
    position: int


class ExcelCellBatchRequest(BaseModel):
    sheet: str
    cells: list[str] = Field(min_length=1, max_length=500)


class WorkspaceExcelSheetRequest(BaseModel):
    upload_id: str
    sheet: str


class ExcelTableCellStyle(BaseModel):
    textColor: str | None = None
    backgroundColor: str | None = None


class WorkspaceWidget(BaseModel):
    id: str
    type: str
    label: str
    slot: int
    width: int = 1
    height: int = 1
    content: str | None = None
    fontSize: int = 16
    fontFamily: str = "Inter"
    fontWeight: int = 700
    fontStyle: str = "normal"
    textColor: str = "#122033"
    backgroundColor: str = "#ffffff"
    useBackgroundColor: bool = False
    textAlign: str = "center"
    verticalAlign: str = "center"
    sourceUploadId: str | None = None
    sourceFilename: str | None = None
    sheetName: str | None = None
    cellRange: str | None = None
    textQueryCell: str | None = None
    tableFontFamily: str | None = None
    tableFontSize: int | None = None
    tableTextAlign: str = "left"
    tableVerticalAlign: str = "middle"
    tableColumnWidths: list[float] = Field(default_factory=list)
    tableRowHeights: list[float] = Field(default_factory=list)
    tableCellStyles: dict[str, ExcelTableCellStyle] = Field(default_factory=dict)
    barItems: list[BarDataItem] = Field(default_factory=list)
    barMax: float = 100
    barMaxInput: str | None = None
    barDisplayPercentage: bool = False
    barMarkerColor: str = "#e42f44"
    barMarkerHeight: int = 5
    barMarkerWidth: int = 3
    barMarkerShowValue: bool = True
    barMarkerFontSize: int = 10
    barBorderRadius: int = 6
    stackCategories: list[StackBarCategory] = Field(default_factory=list)
    stackSeries: list[StackBarSeries] = Field(default_factory=list)
    chartShowLegend: bool = True
    chartLegendPosition: str = "bottom"
    chartColors: list[str] = Field(default_factory=list)
    chartFontSize: int = 10
    pieShowValueCallouts: bool = False
    iconName: str = "package"
    iconColor: str = "#1473e6"
    iconSize: int = 32
    gradientStartColor: str = "#0080c6"
    gradientEndColor: str = "#e42f44"
    gradientDirection: str = "to right"
    gradientStartPosition: int = 0
    gradientEndPosition: int = 100
    gradientOpacity: int = 100
    gradientBorderRadius: int = 8
    gradientStops: list[GradientStop] = Field(default_factory=list)
    lineCurve: str = "smooth"
    lineStrokeWidth: int = 3
    lineShowMarkers: bool = True
    lineNullMissing: bool = False
    lineAnnotations: list[LineAnnotation] = Field(default_factory=list)
    columnLabelRotation: int = -45
    columnWidth: int = 55
    columnBorderRadius: int = 6
    columnShowDataLabels: bool = True
    columnShowDataLabelBackground: bool = True


class WorkspaceBox(BaseModel):
    id: str
    size: str = "1x1"
    title: str = "Box 1x1"
    cell: int
    columns: int
    rows: int
    widgets: list[WorkspaceWidget] = Field(default_factory=list)


class WorkspacePage(BaseModel):
    id: str
    name: str
    boxes: list[WorkspaceBox] = Field(default_factory=list)
    isMain: bool = False
    locked: bool = False


class WorkspaceLayout(BaseModel):
    boxes: list[WorkspaceBox] = Field(default_factory=list)
    pages: list[WorkspacePage] = Field(default_factory=list)
    activePageId: str | None = None


class Base(DeclarativeBase):
    pass


class AppSetting(Base):
    __tablename__ = "app_settings"

    setting_key: Mapped[str] = mapped_column(String(80), primary_key=True)
    setting_value: Mapped[dict[str, Any]] = mapped_column(JSON)
    updated_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now(), onupdate=func.now())


class ExcelUpload(Base):
    __tablename__ = "excel_uploads"

    id: Mapped[str] = mapped_column(String(36), primary_key=True)
    category: Mapped[str] = mapped_column(String(20), nullable=False, default="inbound", server_default="inbound")
    original_filename: Mapped[str] = mapped_column(String(255), nullable=False)
    stored_filename: Mapped[str] = mapped_column(String(255), nullable=False, unique=True)
    content_type: Mapped[str] = mapped_column(String(120), nullable=False)
    file_size: Mapped[int] = mapped_column(BigInteger, nullable=False)
    file_data: Mapped[bytes] = mapped_column(LargeBinary(length=2**32 - 1), nullable=False)
    parsed_data: Mapped[dict[str, Any]] = mapped_column(JSON, nullable=False)
    uploaded_at: Mapped[datetime] = mapped_column(DateTime, server_default=func.now(), nullable=False)


cached_dashboard: dict[str, Any] = {
    "status": "waiting",
    "data": generate_mock_dashboard(),
    "updated_at": "",
    "system_id": SYSTEM_ID,
}

dashboard_settings = DashboardSettings()
WIDGET_LAYOUT_PATH = Path(os.getenv("WIDGET_LAYOUT_PATH", "dashboard_widgets.json"))

DEFAULT_WIDGET_LAYOUT = {
    "widgets": [
        {
            "id": "inbound-widget",
            "title": "Inbound",
            "size": "1x1",
            "items": [
                {"id": "item-inbound-progress", "type": "metric-inbound", "slot": 0, "label": "Inbound Progress"},
            ],
        },
        {
            "id": "pick-widget",
            "title": "Pick",
            "size": "1x1",
            "items": [
                {"id": "item-pick-progress", "type": "metric-pick", "slot": 0, "label": "Pick Progress"},
            ],
        },
        {
            "id": "outbound-widget",
            "title": "Outbound",
            "size": "1x1",
            "items": [
                {"id": "item-outbound-progress", "type": "metric-outbound", "slot": 0, "label": "Outbound Progress"},
            ],
        },
        {
            "id": "progress-chart-widget",
            "title": "Progress Chart",
            "size": "2x1",
            "items": [
                {"id": "item-progress-chart", "type": "progress-chart", "slot": 0, "label": "Progress Chart"},
            ],
        },
    ]
}

DEFAULT_WORKSPACE_LAYOUT = {
    "boxes": [
        {
            "id": "box-1",
            "size": "1x1",
            "title": "Box 1x1",
            "cell": 0,
            "columns": 1,
            "rows": 1,
            "widgets": [
                {"id": "widget-1", "type": "title", "label": "Title", "slot": 0},
            ],
        }
    ]
}

SETTING_KEY_DASHBOARD = "dashboard_settings"
SETTING_KEY_WIDGETS = "dashboard_widgets"
SETTING_KEY_WORKSPACE = "workspace_layout"
SETTING_KEY_FILE_SERVER_SYNC = "file_server_excel_sync"


def build_database_url() -> str:
    if DATABASE_URL:
        return DATABASE_URL
    password = quote_plus(MYSQL_PASSWORD)
    return f"mysql+pymysql://{MYSQL_USER}:{password}@{MYSQL_HOST}:{MYSQL_PORT}/{MYSQL_DATABASE}?charset=utf8mb4"


engine = create_engine(build_database_url(), pool_pre_ping=True, future=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)


def init_database() -> None:
    Base.metadata.create_all(bind=engine)
    with engine.begin() as connection:
        category_column = connection.execute(
            text(
                "SELECT COUNT(*) FROM information_schema.COLUMNS "
                "WHERE TABLE_SCHEMA = :database AND TABLE_NAME = 'excel_uploads' AND COLUMN_NAME = 'category'"
            ),
            {"database": MYSQL_DATABASE},
        ).scalar_one()
        if not category_column:
            connection.execute(text("ALTER TABLE excel_uploads ADD COLUMN category VARCHAR(20) NOT NULL DEFAULT 'inbound' AFTER id"))
            connection.execute(text("CREATE INDEX idx_excel_uploads_category ON excel_uploads (category)"))


@app.on_event("startup")
async def startup() -> None:
    global file_server_sync_task
    try:
        init_database()
    except SQLAlchemyError:
        # Keep API bootable while MySQL is being prepared; DB-backed endpoints return 503.
        pass
    if FILE_SERVER_SYNC_ENABLED:
        file_server_sync_task = asyncio.create_task(file_server_sync_loop())


@app.on_event("shutdown")
async def shutdown() -> None:
    global file_server_sync_task
    if file_server_sync_task is None:
        return
    file_server_sync_task.cancel()
    with contextlib.suppress(asyncio.CancelledError):
        await file_server_sync_task
    file_server_sync_task = None


def read_setting(setting_key: str) -> dict[str, Any] | None:
    with SessionLocal() as session:
        setting = session.get(AppSetting, setting_key)
        return setting.setting_value if setting else None


def write_setting(setting_key: str, setting_value: dict[str, Any]) -> dict[str, Any]:
    with SessionLocal() as session:
        setting = session.get(AppSetting, setting_key)
        if setting is None:
            setting = AppSetting(setting_key=setting_key, setting_value=setting_value)
            session.add(setting)
        else:
            setting.setting_value = setting_value
        session.commit()
        return setting_value


def require_database_error(exc: SQLAlchemyError) -> HTTPException:
    return HTTPException(status_code=503, detail=f"Database unavailable: {exc}")


def now_string() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def load_widget_layout() -> dict[str, Any]:
    try:
        return read_setting(SETTING_KEY_WIDGETS) or DEFAULT_WIDGET_LAYOUT
    except SQLAlchemyError:
        pass

    if not WIDGET_LAYOUT_PATH.exists():
        return DEFAULT_WIDGET_LAYOUT
    try:
        with WIDGET_LAYOUT_PATH.open("r", encoding="utf-8") as file:
            return json.load(file)
    except (OSError, json.JSONDecodeError):
        return DEFAULT_WIDGET_LAYOUT


def save_widget_layout(layout: DashboardWidgetLayout) -> dict[str, Any]:
    data = layout.model_dump()
    try:
        return write_setting(SETTING_KEY_WIDGETS, data)
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc


def load_dashboard_settings() -> DashboardSettings:
    try:
        data = read_setting(SETTING_KEY_DASHBOARD)
        if data:
            return DashboardSettings(**data)
    except SQLAlchemyError:
        pass
    return dashboard_settings


def save_dashboard_settings(settings: DashboardSettings) -> dict[str, Any]:
    data = settings.model_dump()
    try:
        return write_setting(SETTING_KEY_DASHBOARD, data)
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc


def load_workspace_layout() -> dict[str, Any]:
    try:
        return read_setting(SETTING_KEY_WORKSPACE) or DEFAULT_WORKSPACE_LAYOUT
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc


def save_workspace_layout(layout: WorkspaceLayout) -> dict[str, Any]:
    try:
        return write_setting(SETTING_KEY_WORKSPACE, layout.model_dump())
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc


def select_latest_workbook_sheet(file_data: bytes) -> str:
    workbook = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
    try:
        dated_sheets: list[tuple[datetime, str]] = []
        for sheet_name in workbook.sheetnames:
            try:
                sheet_date = datetime.strptime(" ".join(sheet_name.split()), "%d %b %Y")
                dated_sheets.append((sheet_date, sheet_name))
            except ValueError:
                continue
        if dated_sheets:
            return max(dated_sheets, key=lambda item: item[0])[1]
        if not workbook.sheetnames:
            raise ValueError("The latest File Server workbook has no sheets")
        return workbook.sheetnames[-1]
    finally:
        workbook.close()


def update_workspace_widgets_sheet(upload_id: str, sheet_name: str, rebind_orphaned: bool = False) -> int:
    layout = WorkspaceLayout(**load_workspace_layout())
    updated_widget_ids: set[str] = set()
    existing_upload_ids: set[str] = set()
    if rebind_orphaned:
        with SessionLocal() as session:
            existing_upload_ids = set(session.scalars(select(ExcelUpload.id)).all())

    def update_boxes(boxes: list[WorkspaceBox]) -> None:
        for box in boxes:
            for widget in box.widgets:
                uses_target = widget.sourceUploadId == upload_id
                uses_missing_upload = bool(
                    rebind_orphaned
                    and widget.sourceUploadId
                    and widget.sourceUploadId not in existing_upload_ids
                )
                if (uses_target or uses_missing_upload) and (
                    widget.sourceUploadId != upload_id or widget.sheetName != sheet_name
                ):
                    widget.sourceUploadId = upload_id
                    widget.sheetName = sheet_name
                    updated_widget_ids.add(widget.id)

    update_boxes(layout.boxes)
    for page in layout.pages:
        update_boxes(page.boxes)
    if updated_widget_ids:
        save_workspace_layout(layout)
    return len(updated_widget_ids)


@app.get("/health")
def health() -> dict[str, str]:
    database_status = "ok"
    try:
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
    except SQLAlchemyError:
        database_status = "unavailable"

    return {"status": "ok", "system_id": SYSTEM_ID, "database": database_status}


@app.get("/data")
def get_dashboard_data() -> dict[str, Any]:
    if USE_MOCK and cached_dashboard["status"] == "waiting":
        cached_dashboard.update(
            {
                "status": "success",
                "data": generate_mock_dashboard(),
                "updated_at": now_string(),
            }
        )

    return {
        "status": cached_dashboard["status"],
        "data": cached_dashboard["data"],
        "updated_at": cached_dashboard["updated_at"],
        "system_id": cached_dashboard["system_id"],
        "settings": load_dashboard_settings().model_dump(),
    }


def serialize_excel_upload(upload: ExcelUpload, managed_upload_id: str | None = None) -> dict[str, Any]:
    return {
        "id": upload.id,
        "category": upload.category,
        "filename": upload.original_filename,
        "content_type": upload.content_type,
        "file_size": upload.file_size,
        "uploaded_at": upload.uploaded_at.isoformat() if upload.uploaded_at else None,
        "managed": upload.id == managed_upload_id,
    }


def normalize_json_numbers(value: Any) -> Any:
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, dict):
        return {key: normalize_json_numbers(item) for key, item in value.items()}
    if isinstance(value, list):
        return [normalize_json_numbers(item) for item in value]
    return value


def normalize_upload_category(category: str) -> str:
    normalized = category.strip().lower()
    if normalized not in {"inbound", "pick", "outbound"}:
        raise HTTPException(status_code=400, detail="Upload category must be inbound, pick, or outbound")
    return normalized


def find_latest_file_server_excel() -> Path:
    source_directory = Path(FILE_SERVER_PATH)
    if not source_directory.is_dir():
        raise FileNotFoundError(f"File Server path is unavailable: {FILE_SERVER_PATH}")
    candidates = [
        path for path in source_directory.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError("No .xlsx files were found on the File Server")
    return max(candidates, key=lambda path: (path.stat().st_mtime_ns, path.name.lower()))


def sync_latest_file_server_excel() -> dict[str, Any]:
    if not FILE_SERVER_SYNC_ENABLED:
        raise RuntimeError("Automatic File Server sync is disabled")
    if not file_server_sync_lock.acquire(blocking=False):
        raise RuntimeError("A File Server sync is already running")

    file_server_sync_status.update({"state": "checking", "last_checked_at": now_string(), "message": "Scanning File Server"})
    try:
        source_path = find_latest_file_server_excel()
        source_stat = source_path.stat()
        previous_state = read_setting(SETTING_KEY_FILE_SERVER_SYNC) or {}
        signature = {
            "source_path": str(source_path),
            "source_mtime_ns": source_stat.st_mtime_ns,
            "source_size": source_stat.st_size,
        }
        upload_id = previous_state.get("upload_id")
        upload_exists = False
        if upload_id:
            with SessionLocal() as session:
                upload_exists = session.get(ExcelUpload, upload_id) is not None
        if upload_id and upload_exists and all(previous_state.get(key) == value for key, value in signature.items()):
            file_server_sync_status.update({
                "state": "idle",
                "latest_filename": source_path.name,
                "upload_id": upload_id,
                "message": "Latest file is already synchronized",
            })
            return {"changed": False, "upload_id": upload_id, "filename": source_path.name}

        file_data = source_path.read_bytes()
        if not file_data:
            raise ValueError("The latest File Server workbook is empty")
        if len(file_data) > FILE_SERVER_MAX_FILE_BYTES:
            raise ValueError(f"The latest File Server workbook exceeds {FILE_SERVER_MAX_FILE_BYTES // (1024 * 1024)} MB")

        upload_id = upload_id or str(uuid.uuid4())
        original_filename = source_path.name
        stored_filename = f"{upload_id}_{original_filename}"
        UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
        target_path = UPLOAD_ROOT / stored_filename
        target_path.write_bytes(file_data)
        try:
            dashboard_data = process_excel_file(target_path)
        except Exception:
            target_path.unlink(missing_ok=True)
            raise
        normalized_data = normalize_json_numbers(jsonable_encoder(dashboard_data))
        latest_sheet = select_latest_workbook_sheet(file_data)

        previous_stored_filename: str | None = None
        with SessionLocal() as session:
            upload = session.get(ExcelUpload, upload_id)
            if upload is None:
                upload = ExcelUpload(
                    id=upload_id,
                    category="inbound",
                    original_filename=original_filename,
                    stored_filename=stored_filename,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    file_size=len(file_data),
                    file_data=file_data,
                    parsed_data=normalized_data,
                )
                session.add(upload)
            else:
                previous_stored_filename = upload.stored_filename
                upload.original_filename = original_filename
                upload.stored_filename = stored_filename
                upload.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                upload.file_size = len(file_data)
                upload.file_data = file_data
                upload.parsed_data = normalized_data
                upload.uploaded_at = datetime.now()
            session.commit()

        if previous_stored_filename and previous_stored_filename != stored_filename:
            (UPLOAD_ROOT / previous_stored_filename).unlink(missing_ok=True)
        sync_state = {
            **signature,
            "upload_id": upload_id,
            "filename": original_filename,
            "synced_at": now_string(),
        }
        write_setting(SETTING_KEY_FILE_SERVER_SYNC, sync_state)
        invalidate_excel_cache(upload_id)
        updated_widgets = update_workspace_widgets_sheet(upload_id, latest_sheet, rebind_orphaned=True)
        cached_dashboard.update({"status": "success", "data": normalized_data, "updated_at": now_string()})
        file_server_sync_status.update({
            "state": "idle",
            "last_synced_at": sync_state["synced_at"],
            "latest_filename": original_filename,
            "upload_id": upload_id,
            "message": "Latest workbook synchronized",
        })
        return {
            "changed": True,
            "upload_id": upload_id,
            "filename": original_filename,
            "sheet": latest_sheet,
            "updated_widgets": updated_widgets,
        }
    except Exception as exc:
        file_server_sync_status.update({"state": "error", "message": str(exc)})
        raise
    finally:
        file_server_sync_lock.release()


async def run_file_server_sync() -> dict[str, Any]:
    if FILE_SERVER_MIRROR_URL:
        def request_mirror() -> None:
            request = UrlRequest(
                FILE_SERVER_MIRROR_URL,
                data=b"",
                headers={"X-Sync-Token": FILE_SERVER_MIRROR_TOKEN},
                method="POST",
            )
            with urlopen(request, timeout=300) as response:
                if response.status >= 400:
                    raise RuntimeError(f"File Server mirror returned HTTP {response.status}")

        await asyncio.to_thread(request_mirror)
    result = await asyncio.to_thread(sync_latest_file_server_excel)
    if result["changed"]:
        if result.get("updated_widgets"):
            broadcast_workspace_event("workspace-layout-saved", datetime.now().isoformat())
        broadcast_workspace_event("excel-upload-replaced", result["upload_id"])
    return result


async def file_server_sync_loop() -> None:
    while True:
        try:
            await run_file_server_sync()
            delay = FILE_SERVER_SYNC_SECONDS
        except Exception:
            delay = min(FILE_SERVER_SYNC_SECONDS, 60)
        await asyncio.sleep(delay)


@app.get("/file-server/status")
def get_file_server_sync_status() -> dict[str, Any]:
    return {"status": "success", "data": dict(file_server_sync_status)}


@app.post("/file-server/sync")
async def sync_file_server_now() -> dict[str, Any]:
    try:
        result = await run_file_server_sync()
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"File Server sync failed: {exc}") from exc
    return {"status": "success", "data": result}


@app.get("/uploads/excel")
def list_excel_uploads() -> dict[str, Any]:
    try:
        managed_upload_id = (read_setting(SETTING_KEY_FILE_SERVER_SYNC) or {}).get("upload_id")
        with SessionLocal() as session:
            uploads = session.scalars(select(ExcelUpload).order_by(ExcelUpload.uploaded_at.desc())).all()
            return {"status": "success", "data": [serialize_excel_upload(upload, managed_upload_id) for upload in uploads]}
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc


def read_excel_upload_data(upload_id: str) -> tuple[str, bytes]:
    try:
        with SessionLocal() as session:
            upload = session.get(ExcelUpload, upload_id)
            if upload is None:
                raise HTTPException(status_code=404, detail="Excel upload not found")
            return upload.original_filename, upload.file_data
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc


def invalidate_excel_cache(upload_id: str) -> None:
    with excel_cache_lock:
        keys = {key for key in excel_sheet_cache if key[0] == upload_id} | {key for key in excel_sheet_locks if key[0] == upload_id}
        locks = [(key, excel_sheet_locks.setdefault(key, Lock())) for key in keys]
    for key, sheet_lock in locks:
        with sheet_lock:
            with excel_cache_lock:
                excel_sheet_cache.pop(key, None)


def get_cached_excel_sheet(upload_id: str, sheet: str) -> tuple[str, list[list[Any]], list[list[str]]]:
    key = (upload_id, sheet)
    now = monotonic()
    with excel_cache_lock:
        cached = excel_sheet_cache.get(key)
        if cached and now - cached[0] < EXCEL_CACHE_TTL_SECONDS:
            excel_sheet_cache.move_to_end(key)
            return cached[1], cached[2], cached[3]
        excel_sheet_cache.pop(key, None)
        sheet_lock = excel_sheet_locks.setdefault(key, Lock())

    with sheet_lock:
        now = monotonic()
        with excel_cache_lock:
            cached = excel_sheet_cache.get(key)
            if cached and now - cached[0] < EXCEL_CACHE_TTL_SECONDS:
                excel_sheet_cache.move_to_end(key)
                return cached[1], cached[2], cached[3]

        filename, file_data = read_excel_upload_data(upload_id)
        try:
            workbook = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
            if sheet not in workbook.sheetnames:
                workbook.close()
                raise HTTPException(status_code=404, detail="Excel sheet not found")
            worksheet = workbook[sheet]
            cells = list(worksheet.iter_rows())
            values = [[cell.value for cell in row] for row in cells]
            number_formats = [[cell.number_format for cell in row] for row in cells]
            workbook.close()
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot read Excel workbook: {exc}") from exc

        with excel_cache_lock:
            excel_sheet_cache[key] = (now, filename, values, number_formats)
            excel_sheet_cache.move_to_end(key)
            while len(excel_sheet_cache) > EXCEL_CACHE_MAX_SHEETS:
                excel_sheet_cache.popitem(last=False)
        return filename, values, number_formats


def read_cached_cell(values: list[list[Any]], row: int, column: int) -> Any:
    if row < 1 or column < 1 or row > len(values):
        return None
    row_values = values[row - 1]
    return row_values[column - 1] if column <= len(row_values) else None


@app.get("/uploads/excel/{upload_id}/sheets")
def list_excel_sheets(upload_id: str) -> dict[str, Any]:
    filename, file_data = read_excel_upload_data(upload_id)
    try:
        workbook = load_workbook(BytesIO(file_data), read_only=True, data_only=True)
        sheets = workbook.sheetnames
        workbook.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel workbook: {exc}") from exc
    return {"status": "success", "filename": filename, "data": sheets}


@app.get("/uploads/excel/{upload_id}/range")
def read_excel_range(upload_id: str, sheet: str, cell_range: str) -> dict[str, Any]:
    normalized_range = cell_range.strip().upper()
    try:
        min_column, min_row, max_column, max_row = range_boundaries(normalized_range)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Cell range must use a format such as A1:J7") from exc

    row_count = max_row - min_row + 1
    column_count = max_column - min_column + 1
    if row_count <= 0 or column_count <= 0 or row_count * column_count > 2000:
        raise HTTPException(status_code=400, detail="Cell range must contain between 1 and 2000 cells")

    filename, sheet_values, sheet_formats = get_cached_excel_sheet(upload_id, sheet)
    values = [
        [read_cached_cell(sheet_values, row, column) for column in range(min_column, max_column + 1)]
        for row in range(min_row, max_row + 1)
    ]
    number_formats = [
        [read_cached_cell(sheet_formats, row, column) or "General" for column in range(min_column, max_column + 1)]
        for row in range(min_row, max_row + 1)
    ]

    return {
        "status": "success",
        "filename": filename,
        "sheet": sheet,
        "cell_range": normalized_range,
        "rows": row_count,
        "columns": column_count,
        "data": jsonable_encoder(values),
        "number_formats": number_formats,
    }


@app.post("/uploads/excel/{upload_id}/cells")
def read_excel_cells(upload_id: str, payload: ExcelCellBatchRequest) -> dict[str, Any]:
    filename, sheet_values, sheet_formats = get_cached_excel_sheet(upload_id, payload.sheet)
    result: dict[str, dict[str, Any]] = {}
    for requested_cell in dict.fromkeys(payload.cells):
        normalized_cell = requested_cell.strip().upper().replace("$", "")
        try:
            min_column, min_row, max_column, max_row = range_boundaries(normalized_cell)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid Cell reference: {requested_cell}") from exc
        if min_column != max_column or min_row != max_row:
            raise HTTPException(status_code=400, detail=f"Expected a single Cell reference: {requested_cell}")
        result[normalized_cell] = {
            "value": jsonable_encoder(read_cached_cell(sheet_values, min_row, min_column)),
            "number_format": read_cached_cell(sheet_formats, min_row, min_column) or "General",
        }
    return {"status": "success", "filename": filename, "sheet": payload.sheet, "data": result}


@app.post("/upload/excel")
async def upload_excel(file: UploadFile = File(...), category: str = Form("inbound")) -> dict[str, Any]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    file_data = await file.read()
    if not file_data:
        raise HTTPException(status_code=400, detail="The uploaded file is empty")
    if len(file_data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Excel file must not exceed 25 MB")
    normalized_category = normalize_upload_category(category)

    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    upload_id = str(uuid.uuid4())
    original_filename = Path(file.filename).name
    stored_filename = f"{upload_id}_{original_filename}"
    target_path = UPLOAD_ROOT / stored_filename
    target_path.write_bytes(file_data)

    try:
        dashboard_data = process_excel_file(target_path)
    except Exception as exc:
        target_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f"Excel parse failed: {exc}") from exc

    normalized_data = normalize_json_numbers(jsonable_encoder(dashboard_data))
    try:
        with SessionLocal() as session:
            upload = ExcelUpload(
                id=upload_id,
                category=normalized_category,
                original_filename=original_filename,
                stored_filename=stored_filename,
                content_type=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                file_size=len(file_data),
                file_data=file_data,
                parsed_data=normalized_data,
            )
            session.add(upload)
            session.commit()
            session.refresh(upload)
            upload_data = serialize_excel_upload(upload)
    except SQLAlchemyError as exc:
        target_path.unlink(missing_ok=True)
        raise require_database_error(exc) from exc

    cached_dashboard.update(
        {
            "status": "success",
            "data": normalized_data,
            "updated_at": now_string(),
        }
    )

    return {
        "status": "success",
        "message": "Excel uploaded and processed",
        "filename": file.filename,
        "stored_path": str(target_path),
        "upload": upload_data,
        "data": normalized_data,
        "updated_at": cached_dashboard["updated_at"],
    }


@app.put("/uploads/excel/{upload_id}")
async def replace_excel_upload(upload_id: str, file: UploadFile = File(...)) -> dict[str, Any]:
    if upload_id == (read_setting(SETTING_KEY_FILE_SERVER_SYNC) or {}).get("upload_id"):
        raise HTTPException(status_code=409, detail="The managed File Server upload cannot be replaced manually")
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    file_data = await file.read()
    if not file_data:
        raise HTTPException(status_code=400, detail="The uploaded file is empty")
    if len(file_data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Excel file must not exceed 25 MB")

    original_filename = Path(file.filename).name
    stored_filename = f"{upload_id}_{original_filename}"
    target_path = UPLOAD_ROOT / stored_filename
    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(file_data)
    try:
        dashboard_data = process_excel_file(target_path)
    except Exception as exc:
        target_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f"Excel parse failed: {exc}") from exc

    normalized_data = normalize_json_numbers(jsonable_encoder(dashboard_data))
    try:
        with SessionLocal() as session:
            upload = session.get(ExcelUpload, upload_id)
            if upload is None:
                target_path.unlink(missing_ok=True)
                raise HTTPException(status_code=404, detail="Excel upload not found")
            previous_stored_filename = upload.stored_filename
            upload.original_filename = original_filename
            upload.stored_filename = stored_filename
            upload.content_type = file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            upload.file_size = len(file_data)
            upload.file_data = file_data
            upload.parsed_data = normalized_data
            upload.uploaded_at = datetime.now()
            session.commit()
            session.refresh(upload)
            upload_data = serialize_excel_upload(upload)
    except SQLAlchemyError as exc:
        target_path.unlink(missing_ok=True)
        raise require_database_error(exc) from exc

    if previous_stored_filename != stored_filename:
        (UPLOAD_ROOT / previous_stored_filename).unlink(missing_ok=True)
    invalidate_excel_cache(upload_id)
    cached_dashboard.update({"status": "success", "data": normalized_data, "updated_at": now_string()})
    broadcast_workspace_event("excel-upload-replaced", upload_id)
    return {
        "status": "success",
        "message": "Excel upload replaced",
        "upload": upload_data,
        "data": normalized_data,
        "updated_at": cached_dashboard["updated_at"],
    }


@app.delete("/uploads/excel/{upload_id}")
def delete_excel_upload(upload_id: str) -> dict[str, str]:
    if upload_id == (read_setting(SETTING_KEY_FILE_SERVER_SYNC) or {}).get("upload_id"):
        raise HTTPException(status_code=409, detail="The managed File Server upload cannot be deleted")
    try:
        with SessionLocal() as session:
            upload = session.get(ExcelUpload, upload_id)
            if upload is None:
                raise HTTPException(status_code=404, detail="Excel upload not found")
            stored_filename = upload.stored_filename
            session.delete(upload)
            session.commit()
    except SQLAlchemyError as exc:
        raise require_database_error(exc) from exc

    (UPLOAD_ROOT / stored_filename).unlink(missing_ok=True)
    invalidate_excel_cache(upload_id)
    return {"status": "success", "message": "Excel upload deleted"}


@app.post("/auth/login")
def login(payload: LoginRequest) -> dict[str, str]:
    if not payload.username.strip() or not payload.password.strip():
        raise HTTPException(status_code=400, detail="Username and password are required")

    return {
        "status": "success",
        "access_token": "dev-token",
        "token_type": "bearer",
    }


@app.get("/settings/dashboard")
def get_settings() -> dict[str, Any]:
    return {"status": "success", "data": load_dashboard_settings().model_dump()}


@app.put("/settings/dashboard")
def update_settings(settings: DashboardSettings) -> dict[str, Any]:
    global dashboard_settings
    dashboard_settings = settings
    return {"status": "success", "data": save_dashboard_settings(settings)}


@app.get("/dashboard/widgets")
def get_dashboard_widgets() -> dict[str, Any]:
    return {"status": "success", "data": load_widget_layout()}


@app.put("/dashboard/widgets")
def update_dashboard_widgets(layout: DashboardWidgetLayout) -> dict[str, Any]:
    return {"status": "success", "data": save_widget_layout(layout)}


@app.get("/workspace/layout")
def get_workspace_layout() -> dict[str, Any]:
    return {"status": "success", "data": load_workspace_layout()}


@app.get("/workspace/events")
async def workspace_events(request: Request) -> StreamingResponse:
    event_queue: asyncio.Queue[tuple[str, str]] = asyncio.Queue(maxsize=10)
    workspace_event_subscribers.add(event_queue)

    async def stream():
        try:
            yield "event: connected\ndata: ready\n\n"
            while not await request.is_disconnected():
                try:
                    event_name, event_data = await asyncio.wait_for(event_queue.get(), timeout=20)
                    yield f"event: {event_name}\ndata: {event_data}\n\n"
                except TimeoutError:
                    yield ": heartbeat\n\n"
        finally:
            workspace_event_subscribers.discard(event_queue)

    return StreamingResponse(
        stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.put("/workspace/layout")
async def update_workspace_layout(layout: WorkspaceLayout) -> dict[str, Any]:
    saved_layout = save_workspace_layout(layout)
    broadcast_workspace_event("workspace-layout-saved", datetime.now().isoformat())
    return {"status": "success", "data": saved_layout}


@app.put("/workspace/excel-sheet")
def update_workspace_excel_sheet(payload: WorkspaceExcelSheetRequest) -> dict[str, Any]:
    sheet_name = payload.sheet
    if not sheet_name.strip():
        raise HTTPException(status_code=400, detail="Sheet is required")
    get_cached_excel_sheet(payload.upload_id, sheet_name)

    sync_state = read_setting(SETTING_KEY_FILE_SERVER_SYNC) or {}
    is_managed_upload = sync_state.get("upload_id") == payload.upload_id
    updated_widgets = update_workspace_widgets_sheet(
        payload.upload_id,
        sheet_name,
        rebind_orphaned=is_managed_upload,
    )
    if not updated_widgets:
        raise HTTPException(status_code=404, detail="No workspace widgets use this Excel file")

    broadcast_workspace_event("workspace-layout-saved", datetime.now().isoformat())
    return {
        "status": "success",
        "sheet": sheet_name,
        "updated_widgets": updated_widgets,
        "data": load_workspace_layout(),
    }
